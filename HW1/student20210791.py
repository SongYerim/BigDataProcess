#!/usr/bin/python3
from openpyxl import load_workbook

wb = load_workbook(filename='student.xlsx')
ws = wb['Sheet1']

r_max = ws.max_row

for r in range(2,r_max+1):
  ws['G'+str(r)].value = (ws['C'+str(r)].value*0.3) + (ws['D'+str(r)].value*0.35) + (ws['E'+str(r)].value*0.34) + (ws['F'+str(r)].value)

total=[]
for r in range(2,r_max+1):
  total.append(ws.cell(row=r,column=7).value)
desc = total[:]
desc.sort(reverse=True)

stu_cnt=len(total)
a0 = int(stu_cnt*0.3)-1
b0 = int(stu_cnt*0.7)-1

if desc[a0] == desc[a0+1]:
  j = a0-1
  for i in range(a0):
    if desc[j] == desc[j+1]:
      j -= 1
      continue
    else:
      a0 = j
if desc[b0] == desc[b0+1]:
  j = b0-1
  for i in range(b0-a0):
    if desc[j] == desc[j+1]:
      j -= 1
      continue
    else:
      b0 = j

list_a = []
for i in range(0,a0+1):
  list_a.append(desc[i])
list_b = []
for i in range(a0+1,b0+1):
  list_b.append(desc[i])

cut_a = (len(list_a)-1)//2
cut_b = (len(list_a)-1) + len(list_b)//2
cut_c = b0 + (len(desc)-b0-1)//2

if desc[cut_a] == desc[cut_a+1]:
  j = cut_a-1
  for i in range(cut_a):
    if desc[j] == desc[j+1]:
      j -= 1
      continue
    else:
      cut_a = j
if desc[cut_b] == desc[cut_b+1]:
  j = cut_b-1
  for i in range(cut_b-a0):
    if desc[j] == desc[j+1]:
      j -= 1
      continue
    else:
      cut_b = j
if desc[cut_c] == desc[cut_c+1]:
  j = cut_c-1
  for i in range(cut_c-b0):
    if desc[j] == desc[j+1]:
      j -= 1
      continue
    else:
      cut_c = j
      
for r in range(2,r_max+1):
  tmp = ws.cell(row=r, column=7).value
  if tmp >= desc[cut_a]:
    ws.cell(row=r,column=8,value='A+')
  elif tmp >= desc[a0]:
    ws.cell(row=r,column=8,value='A0')
  elif tmp >= desc[cut_b]:
    ws.cell(row=r,column=8,value='B+')
  elif tmp >= desc[b0]:
    ws.cell(row=r,column=8,value='B0')
  elif tmp >= desc[cut_c]:
    ws.cell(row=r,column=8,value='C+')
  else:
    ws.cell(row=r,column=8,value='C0')
  if tmp < 40:
    ws.cell(row=r,column=8,value='F')
    
wb.save(filename='student.xlsx')
